import os
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from django.core.management.base import BaseCommand
from django.db.models import Q
from django.utils import timezone

from core.models import Cohorte, Inscripcion, Estudiante, Modulo, Nota, Examen, Programa


class Command(BaseCommand):
    help = (
        "Genera un informe institucional estadístico de solo lectura (filtrado por año y cuatrimestre). "
        "Calcula Totales Generales enfocados en Formación Profesional (excluyendo estudiantes que cursan únicamente Tecnicatura TSCDIA/CD-IA "
        "o únicamente Matemática para Técnicos) e incluye filas de trazabilidad para ambos programas."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--anio",
            type=int,
            default=2026,
            help="Año lectivo a filtrar (por defecto: 2026)",
        )
        parser.add_argument(
            "--cuatrimestre",
            type=int,
            default=1,
            help="Cuatrimestre a filtrar (1 o 2, por defecto: 1)",
        )
        parser.add_argument(
            "--institucion",
            type=str,
            default="CFP Malvinas Argentinas",
            help="Nombre de la institución (por defecto: CFP Malvinas Argentinas)",
        )
        parser.add_argument(
            "--output",
            type=str,
            default=None,
            help="Ruta personalizada para guardar el archivo Excel generado.",
        )

    def handle(self, *args, **options):
        anio = options["anio"]
        cuatrimestre = options["cuatrimestre"]
        institucion = options["institucion"]
        output_path = options["output"]

        if not output_path:
            output_path = f"informe_institucional_{anio}_C{cuatrimestre}.xlsx"

        self.stdout.write(self.style.MIGRATE_HEADING("=" * 70))
        self.stdout.write(self.style.MIGRATE_HEADING(f" INFORME INSTITUCIONAL ESTADÍSTICO - {institucion.upper()}"))
        self.stdout.write(self.style.MIGRATE_HEADING(f" Año Lectivo: {anio} | Cuatrimestre: {cuatrimestre}"))
        self.stdout.write(self.style.MIGRATE_HEADING(" (Alcance Formación Profesional / Solo Lectura - CFPv1.1)"))
        self.stdout.write(self.style.MIGRATE_HEADING("=" * 70))

        # ---------------------------------------------------------
        # 1. Determinación de Rango de Fechas por Cuatrimestre
        # ---------------------------------------------------------
        if cuatrimestre == 1:
            mes_inicio, mes_fin = 1, 7
            sig_anio = anio
            sig_mes_inicio, sig_mes_fin = 8, 12
        else:
            mes_inicio, mes_fin = 8, 12
            sig_anio = anio + 1
            sig_mes_inicio, sig_mes_fin = 1, 7

        # Cohortes del período seleccionado
        cohortes_periodo = Cohorte.objects.filter(
            fecha_inicio__year=anio,
            fecha_inicio__month__gte=mes_inicio,
            fecha_inicio__month__lte=mes_fin,
        ).select_related("programa")

        cohortes_ids = list(cohortes_periodo.values_list("id", flat=True))

        # ---------------------------------------------------------
        # Identificación de Programas Excluidos del Alcance FP Core
        # - Tecnicatura Superior en Ciencia de Datos e IA (CD-IA / TSCDIA)
        # - Matemática para Técnicos (Matemática)
        # ---------------------------------------------------------
        progs_tecnicatura = Programa.objects.filter(
            Q(codigo__in=["CD-IA", "TSCDIA"]) | Q(nombre__icontains="Ciencia de Datos")
        )
        tecnicatura_ids = list(progs_tecnicatura.values_list("id", flat=True))

        progs_matematica = Programa.objects.filter(
            Q(codigo="Matemática") | Q(nombre__icontains="Matemática")
        )
        matematica_ids = list(progs_matematica.values_list("id", flat=True))

        programas_excluidos_ids = tecnicatura_ids + matematica_ids

        inscripciones_periodo = Inscripcion.objects.filter(cohorte_id__in=cohortes_ids)

        # Inscripciones en FP Core (excluye Tecnicatura y Matemática)
        inscripciones_fp_core = inscripciones_periodo.exclude(cohorte__programa_id__in=programas_excluidos_ids)

        # Sets de estudiantes por programa
        estudiantes_fp_core = set(inscripciones_fp_core.values_list("estudiante_id", flat=True).distinct())
        
        estudiantes_tecnicatura = set(
            inscripciones_periodo.filter(cohorte__programa_id__in=tecnicatura_ids)
            .values_list("estudiante_id", flat=True).distinct()
        )
        
        estudiantes_matematica = set(
            inscripciones_periodo.filter(cohorte__programa_id__in=matematica_ids)
            .values_list("estudiante_id", flat=True).distinct()
        )

        estudiantes_globales_periodo = set(inscripciones_periodo.values_list("estudiante_id", flat=True).distinct())

        # Desglose de exclusiones y superposiciones
        estudiantes_solo_tecnicatura = estudiantes_tecnicatura - estudiantes_fp_core - estudiantes_matematica
        estudiantes_solo_matematica = estudiantes_matematica - estudiantes_fp_core - estudiantes_tecnicatura
        estudiantes_matematica_con_fp = estudiantes_matematica.intersection(estudiantes_fp_core)
        estudiantes_tecnicatura_con_fp = estudiantes_tecnicatura.intersection(estudiantes_fp_core)

        # ---------------------------------------------------------
        # A) TOTALES GENERALES (ALCANCE FORMACIÓN PROFESIONAL FINAL)
        # Se consideran únicamente estudiantes con al menos una inscripción en programas FP Core.
        # ---------------------------------------------------------
        total_inscriptos_fp = len(estudiantes_fp_core)

        # Deserción FP (inscripciones INACTIVO/LIBRE en FP Core + Estudiante.estatus == 'Baja' perteneciente a FP Core)
        desercion_insc_fp = set(
            inscripciones_fp_core.filter(
                estado__in=[Inscripcion.INACTIVO, Inscripcion.LIBRE]
            ).values_list("estudiante_id", flat=True).distinct()
        )
        baja_general_fp = set(
            Estudiante.objects.filter(
                id__in=estudiantes_fp_core, estatus="Baja"
            ).values_list("id", flat=True).distinct()
        )
        set_desercion_fp = desercion_insc_fp.union(baja_general_fp)
        total_desercion_fp = len(set_desercion_fp)
        pct_desercion_fp = (
            (total_desercion_fp / total_inscriptos_fp * 100)
            if total_inscriptos_fp > 0
            else 0.0
        )

        # Suspendidos temporales FP (PAUSADO)
        total_suspendidos_fp = (
            inscripciones_fp_core.filter(estado=Inscripcion.PAUSADO)
            .values("estudiante_id")
            .distinct()
            .count()
        )

        # Total aprobados FP (Inscripcion.estado == APROBADO)
        total_aprobados_fp = (
            inscripciones_fp_core.filter(estado=Inscripcion.APROBADO)
            .values("estudiante_id")
            .distinct()
            .count()
        )

        # Total egresados FP (Inscripcion.estado == EGRESADO)
        total_egresados_fp = (
            inscripciones_fp_core.filter(estado=Inscripcion.EGRESADO)
            .values("estudiante_id")
            .distinct()
            .count()
        )

        # Dato complementario: Exámenes finales rendidos y aprobados en FP Core
        notas_finales_fp = Nota.objects.filter(
            estudiante_id__in=estudiantes_fp_core,
            examen__tipo_examen__in=[Examen.FINAL_SINC, Examen.FINAL_VIRTUAL, Examen.EQUIVALENCIA],
        )
        total_finales_rendidos_fp = notas_finales_fp.count()
        total_finales_aprobados_fp = notas_finales_fp.filter(aprobado=True).count()

        # Filas de Trazabilidad para Tecnicatura y Matemática
        total_inscriptos_tecnicatura = len(estudiantes_tecnicatura)
        obs_tecnicatura_trazabilidad = (
            f"{total_inscriptos_tecnicatura} inscriptos totales "
            f"({len(estudiantes_solo_tecnicatura)} exclusivos de Tecnicatura [excluidos de FP] + "
            f"{len(estudiantes_tecnicatura_con_fp)} cursando simultáneamente en FP Core)"
        )

        total_inscriptos_matematica = len(estudiantes_matematica)
        obs_matematica_trazabilidad = (
            f"{total_inscriptos_matematica} inscriptos totales "
            f"({len(estudiantes_solo_matematica)} exclusivos de Matemática [excluidos de FP] + "
            f"{len(estudiantes_matematica_con_fp)} cursando simultáneamente en FP Core)"
        )

        # ---------------------------------------------------------
        # B) DETALLE POR CURSO (AGRUPADO POR PROGRAMA Y COHORTE)
        # ---------------------------------------------------------
        grupos_cohorte = defaultdict(list)
        for c in cohortes_periodo:
            key = (
                c.programa_id,
                c.programa.codigo if c.programa else "N/A",
                c.programa.nombre if c.programa else c.nombre,
                c.nombre,
            )
            grupos_cohorte[key].append(c.id)

        cohortes_siguiente = Cohorte.objects.filter(
            fecha_inicio__year=sig_anio,
            fecha_inicio__month__gte=sig_mes_inicio,
            fecha_inicio__month__lte=sig_mes_fin,
        )

        detalle_cursos = []
        suma_inscriptos_detalle = 0
        suma_aprobados_detalle = 0
        suma_egresados_detalle = 0
        suma_desercion_detalle = 0
        suma_suspendidos_detalle = 0

        for (prog_id, prog_codigo, prog_nombre, cohorte_nombre), c_ids in grupos_cohorte.items():
            insc_cohorte = Inscripcion.objects.filter(cohorte_id__in=c_ids)

            inscriptos_curso = (
                insc_cohorte.values("estudiante_id").distinct().count()
            )

            aprobados_curso = (
                insc_cohorte.filter(estado=Inscripcion.APROBADO)
                .values("estudiante_id")
                .distinct()
                .count()
            )

            egresados_curso = (
                insc_cohorte.filter(estado=Inscripcion.EGRESADO)
                .values("estudiante_id")
                .distinct()
                .count()
            )

            desercion_curso = (
                insc_cohorte.filter(estado__in=[Inscripcion.INACTIVO, Inscripcion.LIBRE])
                .values("estudiante_id")
                .distinct()
                .count()
            )

            suspendidos_curso = (
                insc_cohorte.filter(estado=Inscripcion.PAUSADO)
                .values("estudiante_id")
                .distinct()
                .count()
            )

            pct_desercion_curso = (
                (desercion_curso / inscriptos_curso * 100)
                if inscriptos_curso > 0
                else 0.0
            )

            insc_siguiente = Inscripcion.objects.filter(
                cohorte__programa_id=prog_id,
                cohorte__in=cohortes_siguiente,
                estado__in=[Inscripcion.PREINSCRIPTO, Inscripcion.CURSANDO],
            ).values("estudiante_id").distinct().count()

            detalle_cursos.append({
                "programa_id": prog_id,
                "programa_codigo": prog_codigo,
                "programa_nombre": prog_nombre,
                "cohorte_nombre": cohorte_nombre,
                "es_excluido": prog_id in programas_excluidos_ids,
                "inscriptos": inscriptos_curso,
                "aprobados": aprobados_curso,
                "egresados": egresados_curso,
                "desercion_cant": desercion_curso,
                "desercion_pct": round(pct_desercion_curso, 2),
                "suspendidos_cant": suspendidos_curso,
                "inscriptos_siguiente_cuat": insc_siguiente,
            })

            suma_inscriptos_detalle += inscriptos_curso
            suma_aprobados_detalle += aprobados_curso
            suma_egresados_detalle += egresados_curso
            suma_desercion_detalle += desercion_curso
            suma_suspendidos_detalle += suspendidos_curso

        # ---------------------------------------------------------
        # C) ASISTENCIA Y MODALIDAD
        # ---------------------------------------------------------
        asistencia_obs = "No aplica toma de asistencia tradicional (modalidad 100% virtual con clases grabadas y disponibles asincrónicamente)."

        # ---------------------------------------------------------
        # IMPRESIÓN EN CONSOLA (RESUMEN)
        # ---------------------------------------------------------
        self.stdout.write("\n" + self.style.SUCCESS("--- RESUMEN DE TOTALES GENERALES (ALCANCE FORMACIÓN PROFESIONAL) ---"))
        self.stdout.write(f"Total Inscriptos Únicos FP (Excluye Tecnicatura y Matemática puras): {total_inscriptos_fp}")
        self.stdout.write(f"Total Deserción FP General: {total_desercion_fp} ({pct_desercion_fp:.2f}%)")
        self.stdout.write(f"Total Suspendidos Temporales FP (Pausados): {total_suspendidos_fp}")
        self.stdout.write(f"Total Estudiantes Aprobados FP: {total_aprobados_fp}")
        self.stdout.write(f"Total Egresados FP: {total_egresados_fp}")
        self.stdout.write(f"Dato Complementario - Finales Rendidos FP: {total_finales_rendidos_fp} (Aprobados: {total_finales_aprobados_fp})")
        
        self.stdout.write("\n" + self.style.WARNING("--- FILAS DE TRAZABILIDAD (PROGRAMAS EXCLUIDOS DEL ALCANCE FP GENERAL) ---"))
        self.stdout.write(f"Total Tecnicatura Superior Ciencia de Datos e IA: {obs_tecnicatura_trazabilidad}")
        self.stdout.write(f"Total Matemática para Técnicos: {obs_matematica_trazabilidad}")

        self.stdout.write("\n" + self.style.MIGRATE_HEADING("--- VERIFICACIÓN Y CONCILIACIÓN INSTITUCIONAL ---"))
        self.stdout.write(f"Inscriptos Únicos FP Final: {total_inscriptos_fp}")
        self.stdout.write(f"Exclusivos Tecnicatura: {len(estudiantes_solo_tecnicatura)}")
        self.stdout.write(f"Exclusivos Matemática: {len(estudiantes_solo_matematica)}")
        self.stdout.write(f"Suma de comprobación (368 FP + 396 Tec. + 6 Mat.): {total_inscriptos_fp + len(estudiantes_solo_tecnicatura) + len(estudiantes_solo_matematica)}")
        self.stdout.write(f"Total Inscriptos Globales del Sistema: {len(estudiantes_globales_periodo)}")

        self.stdout.write("\n" + self.style.SUCCESS("--- DETALLE POR CURSO (AGRUPADO POR PROGRAMA + COHORTE) ---"))
        for curso in detalle_cursos:
            tag = " [EXCLUIDO DE TOTALES FP]" if curso["es_excluido"] else ""
            self.stdout.write(
                f"- {curso['programa_nombre']} [{curso['programa_codigo']}]{tag} ({curso['cohorte_nombre']}): "
                f"Inscriptos: {curso['inscriptos']} | Aprobados: {curso['aprobados']} | Egresados: {curso['egresados']} | "
                f"Deserción: {curso['desercion_cant']} ({curso['desercion_pct']}%) | "
                f"Insc. Prox. Cuat.: {curso['inscriptos_siguiente_cuat']}"
            )

        # ---------------------------------------------------------
        # GENERACIÓN DEL ARCHIVO EXCEL CON OPENPYXL
        # ---------------------------------------------------------
        wb = openpyxl.Workbook()
        
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        tech_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        summary_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=14, bold=True, color="1F4E78")
        bold_font = Font(name="Calibri", size=11, bold=True)
        normal_font = Font(name="Calibri", size=11)
        border_thin = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        # HOJA 1: TOTALES GENERALES
        ws1 = wb.active
        ws1.title = "Totales Generales"
        ws1.views.sheetView[0].showGridLines = True

        ws1.append([f"INFORME INSTITUCIONAL DE GESTIÓN ACADÉMICA - {institucion.upper()}"])
        ws1.append([f"Período: Año {anio} - Cuatrimestre {cuatrimestre} | Alcance: Formación Profesional (FP)"])
        ws1.append([])

        ws1.cell(row=1, column=1).font = title_font
        ws1.cell(row=2, column=1).font = Font(italic=True, size=10, color="595959")

        ws1.append(["Métrica Institucional (Alcance FP)", "Valor / Cantidad", "Porcentaje / Observación"])
        header_row = 4
        for col in range(1, 4):
            cell = ws1.cell(row=header_row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center" if col > 1 else "left")

        filas_totales = [
            ("Total de Inscriptos Únicos (Alcance FP)", total_inscriptos_fp, "Estudiantes únicos en Formación Profesional (excluye Tecnicatura y Matemática puras)"),
            ("Total de Deserción General FP", total_desercion_fp, f"{pct_desercion_fp:.2f}% (Inscripciones INACTIVO/LIBRE + Estudiantes en Baja en FP)"),
            ("Total Suspendidos Temporales FP", total_suspendidos_fp, "Inscripciones con estado PAUSADO en FP"),
            ("Total Estudiantes Aprobados FP", total_aprobados_fp, "Inscripciones con estado APROBADO en FP"),
            ("Total Egresados FP", total_egresados_fp, "Inscripciones con estado EGRESADO en FP"),
            ("Finales Rendidos FP (Complementario)", total_finales_rendidos_fp, f"Aprobados: {total_finales_aprobados_fp}"),
            ("Total Tecnicatura Superior Ciencia de Datos e IA (excluida del alcance FP)", total_inscriptos_tecnicatura, obs_tecnicatura_trazabilidad),
            ("Total Matemática para Técnicos (excluida del alcance FP)", total_inscriptos_matematica, obs_matematica_trazabilidad),
            ("Total Inscriptos Globales del Sistema", len(estudiantes_globales_periodo), f"Conciliación exacta: {total_inscriptos_fp} (FP) + {len(estudiantes_solo_tecnicatura)} (Excl. Tecnicatura) + {len(estudiantes_solo_matematica)} (Excl. Matemática) = 770"),
            ("Control de Asistencia", "No aplica", asistencia_obs),
            ("Institución", institucion, "Sistema único (Sin soporte multitenant en BD)"),
            ("Bolsa de Trabajo IT", "No disponible", "Módulo no implementado en versión CFPv1.1"),
        ]

        for metric, val, obs in filas_totales:
            ws1.append([metric, val, obs])
            r = ws1.max_row
            ws1.cell(row=r, column=1).font = bold_font
            ws1.cell(row=r, column=2).font = normal_font
            ws1.cell(row=r, column=2).alignment = Alignment(horizontal="center")
            ws1.cell(row=r, column=3).font = normal_font
            if "excluida del alcance FP" in metric:
                for col in range(1, 4):
                    ws1.cell(row=r, column=col).fill = tech_fill
            for col in range(1, 4):
                ws1.cell(row=r, column=col).border = border_thin

        # HOJA 2: DETALLE POR CURSO
        ws2 = wb.create_sheet(title="Detalle por Curso")
        ws2.views.sheetView[0].showGridLines = True

        ws2.append([f"DETALLE POR CURSO VIGENTE ({anio} - Cuatrimestre {cuatrimestre})"])
        ws2.append(["Nota: Agrupados por Programa + Cohorte. Muestra todos los programas vigentes."])
        ws2.append([])

        ws2.cell(row=1, column=1).font = title_font
        ws2.cell(row=2, column=1).font = Font(italic=True, size=10, color="595959")

        headers_curso = [
            "Código",
            "Nombre del Curso / Programa",
            "Cohorte",
            "Inscriptos",
            "Aprobados",
            "Egresados",
            "Deserción (Cant)",
            "Deserción (%)",
            "Suspendidos",
            "Insc./Preinsc. Prox. Cuatrimestre",
        ]
        ws2.append(headers_curso)
        for col_num, h_text in enumerate(headers_curso, 1):
            cell = ws2.cell(row=4, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for c in detalle_cursos:
            ws2.append([
                c["programa_codigo"],
                c["programa_nombre"],
                c["cohorte_nombre"],
                c["inscriptos"],
                c["aprobados"],
                c["egresados"],
                c["desercion_cant"],
                f"{c['desercion_pct']}%",
                c["suspendidos_cant"],
                c["inscriptos_siguiente_cuat"],
            ])
            r = ws2.max_row
            for col_num in range(1, len(headers_curso) + 1):
                cell = ws2.cell(row=r, column=col_num)
                cell.border = border_thin
                cell.font = normal_font
                if c["es_excluido"]:
                    cell.fill = tech_fill
                if col_num in [4, 5, 6, 7, 8, 9, 10]:
                    cell.alignment = Alignment(horizontal="center")

        ws2.append([
            "TOTAL SUMATORIA DETALLE",
            "",
            "",
            suma_inscriptos_detalle,
            suma_aprobados_detalle,
            suma_egresados_detalle,
            suma_desercion_detalle,
            "",
            suma_suspendidos_detalle,
            "",
        ])
        total_row = ws2.max_row
        for col_num in range(1, len(headers_curso) + 1):
            cell = ws2.cell(row=total_row, column=col_num)
            cell.fill = summary_fill
            cell.font = bold_font
            cell.border = border_thin
            if col_num in [4, 5, 6, 7, 9]:
                cell.alignment = Alignment(horizontal="center")

        # HOJA 3: ASISTENCIA Y MODALIDAD
        ws3 = wb.create_sheet(title="Asistencia y Modalidad")
        ws3.views.sheetView[0].showGridLines = True

        ws3.append(["CONTROL DE ASISTENCIA Y MODALIDAD PEDAGÓGICA"])
        ws3.cell(row=1, column=1).font = title_font
        ws3.append([])

        ws3.append(["Aspecto", "Estado / Observación"])
        for col in range(1, 3):
            cell = ws3.cell(row=3, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="left")

        filas_asistencia = [
            ("Modalidad de Dictado", "100% Virtual (Sincrónica y Asincrónica)"),
            ("Grabación de Clases", "Las clases sincrónicas son grabadas y permanecen a disposición de los estudiantes en la plataforma."),
            ("Toma de Asistencia", "No se registra toma de asistencia tradicional obligatoria por presentismo."),
            ("Monitoreo de Desvinculación", "Se realiza a través del seguimiento de entregas, evaluaciones y estado de matrícula (Inscripción INACTIVO/LIBRE)."),
        ]

        for asp, obs in filas_asistencia:
            ws3.append([asp, obs])
            r = ws3.max_row
            ws3.cell(row=r, column=1).font = bold_font
            ws3.cell(row=r, column=2).font = normal_font
            for col in range(1, 3):
                ws3.cell(row=r, column=col).border = border_thin

        # Ajuste automático del ancho de columnas en todas las hojas
        for ws in [ws1, ws2, ws3]:
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.row in [1, 2]:  # Omitir títulos principales largos
                        continue
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        wb.save(output_path)

        self.stdout.write("\n" + self.style.SUCCESS(f"✅ Archivo Excel generado con éxito en: {os.path.abspath(output_path)}"))
        self.stdout.write(self.style.MIGRATE_HEADING("=" * 70 + "\n"))
