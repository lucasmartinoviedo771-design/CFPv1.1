import os
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from django.core.management.base import BaseCommand
from django.db.models import Min, Count
from django.utils import timezone

from core.models import Estudiante, Inscripcion, Nota


class Command(BaseCommand):
    help = (
        "Exporta un reporte académico completo y profesional de estudiantes en formato Excel (.xlsx), "
        "con datos personales, trayectoria académica, cohortes, módulos, estado de regularidad general, "
        "estado/situación específica en la cohorte/módulo y calificaciones históricas."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--output",
            type=str,
            default=None,
            help="Ruta o nombre personalizado del archivo Excel (.xlsx) a generar.",
        )
        parser.add_argument(
            "--programa-id",
            type=int,
            default=None,
            help="Filtrar por ID de Programa específico.",
        )
        parser.add_argument(
            "--cohorte-id",
            type=int,
            default=None,
            help="Filtrar por ID de Cohorte específica.",
        )
        parser.add_argument(
            "--activos-only",
            action="store_true",
            help="Exportar únicamente estudiantes activos (is_active=True).",
        )

    def handle(self, *args, **options):
        output_path = options.get("output")
        programa_id = options.get("programa_id")
        cohorte_id = options.get("cohorte_id")
        activos_only = options.get("activos_only")

        if not output_path:
            timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"reporte_estudiantes_{timestamp}.xlsx"

        if not output_path.endswith(".xlsx"):
            output_path += ".xlsx"

        self.stdout.write(self.style.MIGRATE_HEADING("=" * 75))
        self.stdout.write(self.style.MIGRATE_HEADING(" EXPORTACIÓN DE REPORTE ACADÉMICO DE ESTUDIANTES (.xlsx)"))
        self.stdout.write(self.style.MIGRATE_HEADING("=" * 75))
        self.stdout.write(f"[*] Destino: {output_path}")

        # ---------------------------------------------------------------------
        # 1. Optimización N+1: Agregación de 1° Fecha e Histórico de Reinscripciones
        # ---------------------------------------------------------------------
        self.stdout.write("[*] Consultando primera fecha histórica y conteo de inscripciones por programa...")
        estadisticas_prog_qs = (
            Inscripcion.objects.values("estudiante_id", "cohorte__programa_id")
            .annotate(
                primera_fecha=Min("created_at"),
                total_inscripciones=Count("id")
            )
        )
        # Diccionarios mapeados: (estudiante_id, programa_id)
        primera_fecha_map = {
            (item["estudiante_id"], item["cohorte__programa_id"]): (
                item["primera_fecha"].date() if item["primera_fecha"] else None
            )
            for item in estadisticas_prog_qs
        }
        total_inscripciones_map = {
            (item["estudiante_id"], item["cohorte__programa_id"]): item["total_inscripciones"]
            for item in estadisticas_prog_qs
        }

        # ---------------------------------------------------------------------
        # 2. Obtención de Inscripciones con relaciones precargadas
        # ---------------------------------------------------------------------
        self.stdout.write("[*] Extrayendo trayectoria académica y estudiantes...")
        inscripciones_qs = (
            Inscripcion.objects.select_related(
                "estudiante",
                "cohorte",
                "cohorte__programa",
                "cohorte__bloque",
                "modulo",
                "modulo__bloque",
            )
            .order_by(
                "estudiante__apellido",
                "estudiante__nombre",
                "cohorte__programa__nombre",
                "cohorte__fecha_inicio",
                "modulo__id",
            )
        )

        if activos_only:
            inscripciones_qs = inscripciones_qs.filter(estudiante__is_active=True)
        if programa_id:
            inscripciones_qs = inscripciones_qs.filter(cohorte__programa_id=programa_id)
        if cohorte_id:
            inscripciones_qs = inscripciones_qs.filter(cohorte_id=cohorte_id)

        # ---------------------------------------------------------------------
        # 3. Optimización N+1: Calificaciones de estudiantes asociadas a módulos / bloques
        # ---------------------------------------------------------------------
        self.stdout.write("[*] Extrayendo historial de notas y evaluaciones...")
        notas_qs = (
            Nota.objects.select_related(
                "examen",
                "examen__modulo",
                "examen__bloque",
            )
            .order_by("fecha_calificacion", "id")
        )

        # Mapa de notas:
        # Clave módulo: (estudiante_id, modulo_id) -> lista de dicts de notas
        # Clave bloque: (estudiante_id, bloque_id) -> lista de dicts de notas
        notas_modulo_map = {}
        notas_bloque_map = {}

        for n in notas_qs:
            fecha_rendida = (
                n.fecha_calificacion.date()
                if n.fecha_calificacion
                else (n.created_at.date() if n.created_at else None)
            )
            calif_val = float(n.calificacion) if n.calificacion is not None else None
            nota_info = {
                "nota": calif_val,
                "aprobado": n.aprobado,
                "fecha": fecha_rendida,
                "tipo_examen": n.examen.tipo_examen if n.examen else "",
                "es_definitiva": n.es_nota_definitiva,
            }

            if n.examen and n.examen.modulo_id:
                key = (n.estudiante_id, n.examen.modulo_id)
                notas_modulo_map.setdefault(key, []).append(nota_info)
            elif n.examen and n.examen.bloque_id:
                key = (n.estudiante_id, n.examen.bloque_id)
                notas_bloque_map.setdefault(key, []).append(nota_info)

        # ---------------------------------------------------------------------
        # 4. Creación del Workbook con openpyxl
        # ---------------------------------------------------------------------
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Estudiantes"
        ws.views.sheetView[0].showGridLines = True

        # Estilos visuales profesionales
        header_fill_personal = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # Azul oscuro
        header_fill_estado = PatternFill(start_color="165B33", end_color="165B33", fill_type="solid")    # Verde institucional
        header_fill_academico = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid") # Azul medio
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        data_font = Font(name="Calibri", size=10, color="000000")
        zebra_fill = PatternFill(start_color="F2F7FA", end_color="F2F7FA", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )
        header_border = Border(
            left=Side(style="thin", color="1B365D"),
            right=Side(style="thin", color="1B365D"),
            top=Side(style="medium", color="1B365D"),
            bottom=Side(style="medium", color="1B365D"),
        )

        align_left = Alignment(horizontal="left", vertical="center")
        align_center = Alignment(horizontal="center", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        headers = [
            # 1. Datos Personales
            "Apellido y Nombres",
            "DNI / Documento",
            "Localidad",
            "Nivel de Estudio",
            "Fecha de Nacimiento",
            "Edad",
            # 2. Estado General y Regularidad del Estudiante
            "Activo en Sistema",
            "Estatus de Regularidad",
            "¿Reinscripto en Programa?",
            # 3. Trayectoria Académica, Cursada y Situación
            "Programa / Carrera",
            "1° Fecha Inscripción Programa",
            "Bloque",
            "Módulo / Materia",
            "Cohorte",
            "Estado / Situación en Cursada",
            "Nota / Calificación",
            "Fecha Calificación",
        ]

        ws.row_dimensions[1].height = 28
        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header_title)
            cell.font = header_font
            if col_num <= 6:
                cell.fill = header_fill_personal
            elif col_num <= 9:
                cell.fill = header_fill_estado
            else:
                cell.fill = header_fill_academico
            cell.alignment = header_align
            cell.border = header_border

        # ---------------------------------------------------------------------
        # 5. Generación de Filas
        # ---------------------------------------------------------------------
        hoy = timezone.now().date()
        row_idx = 2
        total_filas = 0

        # Función auxiliar para calcular la edad exacta
        def calcular_edad(fecha_nac):
            if not fecha_nac:
                return None
            try:
                edad = hoy.year - fecha_nac.year - ((hoy.month, hoy.day) < (fecha_nac.month, fecha_nac.day))
                return max(0, edad)
            except Exception:
                return None

        # Diccionario legible para el estado de la inscripción
        dict_estados_inscripcion = dict(Inscripcion.ESTADOS)

        # Procesar inscripciones
        for ins in inscripciones_qs:
            est = ins.estudiante
            cohorte = ins.cohorte
            programa = cohorte.programa if cohorte else None
            modulo = ins.modulo
            
            # Determinar bloque
            bloque = None
            if modulo and modulo.bloque:
                bloque = modulo.bloque
            elif cohorte and cohorte.bloque:
                bloque = cohorte.bloque

            # 1° Fecha de Inscripción histórica al Programa
            primera_fecha_prog = None
            total_inscr_prog = 1
            if programa:
                primera_fecha_prog = primera_fecha_map.get((est.id, programa.id))
                total_inscr_prog = total_inscripciones_map.get((est.id, programa.id), 1)
            if not primera_fecha_prog:
                primera_fecha_prog = ins.created_at.date() if ins.created_at else None

            # Calcular edad
            edad = calcular_edad(est.fecha_nacimiento)

            # Estado del estudiante en el sistema
            activo_str = "Activo" if est.is_active else "Baja / Inactivo"
            estatus_str = est.estatus or "No especificado"
            reinscripto_str = "Sí" if total_inscr_prog > 1 else "No"

            # Estado de la cursada / inscripción específica
            estado_cursada_raw = ins.estado or "PREINSCRIPTO"
            estado_cursada_display = dict_estados_inscripcion.get(estado_cursada_raw, estado_cursada_raw.capitalize())

            # Obtener notas correspondientes
            notas_candidatas = []
            if modulo:
                notas_candidatas = notas_modulo_map.get((est.id, modulo.id), [])
            elif bloque:
                notas_candidatas = notas_bloque_map.get((est.id, bloque.id), [])

            # Si el estudiante rindió múltiples veces (parcial, recuperatorio, final),
            # o si no rindió ninguna:
            if not notas_candidatas:
                # Fila única sin nota
                notas_a_emitir = [{"nota": None, "fecha": None, "label_vacio": "Sin nota"}]
            else:
                notas_a_emitir = notas_candidatas

            for nota_data in notas_a_emitir:
                fill_current = zebra_fill if (row_idx % 2 == 0) else white_fill
                ws.row_dimensions[row_idx].height = 20

                # 1. Apellido y Nombres
                c1 = ws.cell(row=row_idx, column=1, value=f"{est.apellido}, {est.nombre}".strip(", "))
                c1.alignment = align_left

                # 2. DNI / Documento
                c2 = ws.cell(row=row_idx, column=2, value=str(est.dni or "").strip())
                c2.alignment = align_center

                # 3. Localidad
                c3 = ws.cell(row=row_idx, column=3, value=est.ciudad or "Sin especificar")
                c3.alignment = align_left

                # 4. Nivel de estudio
                c4 = ws.cell(row=row_idx, column=4, value=est.nivel_educativo or "No registrado")
                c4.alignment = align_left

                # 5. Fecha de nacimiento
                c5 = ws.cell(row=row_idx, column=5, value=est.fecha_nacimiento if est.fecha_nacimiento else "")
                c5.alignment = align_center
                if est.fecha_nacimiento:
                    c5.number_format = "YYYY-MM-DD"

                # 6. Edad
                c6 = ws.cell(row=row_idx, column=6, value=edad if edad is not None else "")
                c6.alignment = align_center
                if edad is not None:
                    c6.number_format = "#,##0"

                # 7. Activo en Sistema (is_active)
                c7 = ws.cell(row=row_idx, column=7, value=activo_str)
                c7.alignment = align_center

                # 8. Estatus de Regularidad (Regular, Baja, Condicional, Preinscripto)
                c8 = ws.cell(row=row_idx, column=8, value=estatus_str)
                c8.alignment = align_center

                # 9. ¿Reinscripto en Programa?
                c9 = ws.cell(row=row_idx, column=9, value=reinscripto_str)
                c9.alignment = align_center

                # 10. Programa / Carrera
                c10 = ws.cell(row=row_idx, column=10, value=programa.nombre if programa else "Sin Programa")
                c10.alignment = align_left

                # 11. 1° Fecha de Inscripción al Programa
                c11 = ws.cell(row=row_idx, column=11, value=primera_fecha_prog if primera_fecha_prog else "")
                c11.alignment = align_center
                if primera_fecha_prog:
                    c11.number_format = "YYYY-MM-DD"

                # 12. Bloque
                c12 = ws.cell(row=row_idx, column=12, value=bloque.nombre if bloque else "Sin Bloque")
                c12.alignment = align_left

                # 13. Módulo / Materia
                c13 = ws.cell(row=row_idx, column=13, value=modulo.nombre if modulo else "General / Bloque")
                c13.alignment = align_left

                # 14. Cohorte
                c14 = ws.cell(row=row_idx, column=14, value=cohorte.nombre if cohorte else "Sin Cohorte")
                c14.alignment = align_center

                # 15. Estado / Situación en Cursada (Preinscripto, Cursando, Inactivo, Libre, Pausado, Egresado, Aprobado, Desaprobado)
                c15 = ws.cell(row=row_idx, column=15, value=estado_cursada_display)
                c15.alignment = align_center

                # 16. Nota / Calificación
                nota_val = nota_data.get("nota")
                if nota_val is not None:
                    c16 = ws.cell(row=row_idx, column=16, value=float(nota_val))
                    c16.alignment = align_right
                    c16.number_format = "0.00"
                else:
                    c16 = ws.cell(row=row_idx, column=16, value=nota_data.get("label_vacio", "Sin nota"))
                    c16.alignment = align_center

                # 17. Fecha en que rindió / Calificación
                fecha_rendida = nota_data.get("fecha")
                c17 = ws.cell(row=row_idx, column=17, value=fecha_rendida if fecha_rendida else "")
                c17.alignment = align_center
                if fecha_rendida:
                    c17.number_format = "YYYY-MM-DD"

                # Aplicar estilos de celda (fuente, bordes, zebra)
                for col_c in range(1, 18):
                    cell = ws.cell(row=row_idx, column=col_c)
                    cell.font = data_font
                    cell.border = thin_border
                    cell.fill = fill_current

                row_idx += 1
                total_filas += 1

        # ---------------------------------------------------------------------
        # 6. Configuración de Filtros Automáticos y Auto-ajuste de Ancho
        # ---------------------------------------------------------------------
        last_col_letter = get_column_letter(len(headers))
        last_row = max(row_idx - 1, 1)
        ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

        # Ajuste automático del ancho de cada columna con margen
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = 0
            for cell in col:
                val = cell.value
                if val is not None:
                    if isinstance(val, (datetime, date)):
                        str_val = "YYYY-MM-DD"
                    elif isinstance(val, float):
                        str_val = f"{val:.2f}"
                    else:
                        str_val = str(val)
                    max_len = max(max_len, len(str_val))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # ---------------------------------------------------------------------
        # 7. Guardar archivo
        # ---------------------------------------------------------------------
        wb.save(output_path)
        self.stdout.write(self.style.SUCCESS(f"[✓] Archivo Excel generado exitosamente: {output_path}"))
        self.stdout.write(self.style.SUCCESS(f"[✓] Total de registros procesados: {total_filas}"))
        self.stdout.write(self.style.MIGRATE_HEADING("=" * 75))
