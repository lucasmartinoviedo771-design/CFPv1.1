#!/usr/bin/env python3
"""
Carga masiva de estudiantes desde XLSX con normalizacion y log de errores.

Uso sugerido:
  python scripts/carga_estudiantes_masiva.py --file "temporal/Encabezado de carga masiva.xlsx" --dry-run
  python scripts/carga_estudiantes_masiva.py --file "temporal/Encabezado de carga masiva.xlsx"
"""

import argparse
import json
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

# Django setup
PROJECT_ROOT = Path(__file__).resolve().parent.parent
BACKEND_CANDIDATES = [
    Path("/app"),  # docker compose run -v backend:/app
    PROJECT_ROOT / "backend",  # local repo run
]
BACKEND_PATH = next((p for p in BACKEND_CANDIDATES if (p / "academia").exists()), None)
if BACKEND_PATH is None:
    raise SystemExit("ERROR: no se encontro backend/academia para inicializar Django.")

sys.path.insert(0, str(BACKEND_PATH))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "academia.settings")
import django  # noqa: E402

django.setup()

from django.core.exceptions import ValidationError  # noqa: E402
from django.db import transaction  # noqa: E402
from django.utils import timezone  # noqa: E402

from core.models import Estudiante  # noqa: E402
from core.utils.estudiante_normalization import (  # noqa: E402
    normalize_country_with_other as normalize_country_with_other_util,
    normalize_dni_digits,
    normalize_sexo,
    normalize_spaces,
    to_title_case,
    to_upper,
)

EXPECTED_HEADERS = [
    "Marca temporal",
    "Correo Electrónico",
    "Apellidos",
    "Nombres",
    "DNI",
    "CUIL",
    "Sexo",
    "Fecha de Nacimiento",
    "País de Nacimiento",
    "Nacionalidad",
    "Lugar de Nacimiento",
    "Domicilio Particular",
    "Barrio",
    "Ciudad",
    "Teléfono (2964123456)",
    "Estudios de nivel alcanzado",
    "¿Posee pc en su domicilio?",
    "¿Posee conectividad a internet?",
    "¿Puede asistir a clase con su computadora personal?",
    "¿Trabaja?",
    "Lugar de trabajo",
]

SEXO_MAP = {
    "m": "Masculino",
    "masculino": "Masculino",
    "f": "Femenino",
    "femenino": "Femenino",
    "indistinto": "Otro",
    "otro": "Otro",
}

NIVEL_MAP = {
    "primaria completa": "Primaria Completa",
    "nivel primario completo": "Primaria Completa",
    "secundaria incompleta": "Secundaria Incompleta",
    "secundaria completa": "Secundaria Completa",
    "nivel secundario completo": "Secundaria Completa",
    "terciaria/universitaria incompleta": "Terciaria/Universitaria Incompleta",
    "terciaria universitaria incompleta": "Terciaria/Universitaria Incompleta",
    "terciaria/universitaria completa": "Terciaria/Universitaria Completa",
    "terciaria universitaria completa": "Terciaria/Universitaria Completa",
    "nivel terciario completo": "Terciaria/Universitaria Completa",
    "nivel universitario completo": "Terciaria/Universitaria Completa",
    "terciaria/universitaria": "Terciaria/Universitaria",
    "terciaria universitaria": "Terciaria/Universitaria",
}

TRUE_SET = {"si", "sí", "s", "true", "1", "x", "y", "yes"}
FALSE_SET = {"no", "n", "false", "0"}


class RowIssue(Exception):
    pass


def norm_spaces(value: Any) -> str:
    return normalize_spaces(value)


def norm_email(value: Any) -> str:
    return norm_spaces(value).lower()


def norm_dni(value: Any) -> str:
    digits = normalize_dni_digits(value)
    if len(digits) != 8:
        raise RowIssue(f"DNI invalido '{value}'. Debe tener 8 digitos sin puntos.")
    return digits


def norm_cuit(value: Any) -> str:
    return re.sub(r"\s+", "", norm_spaces(value))


def norm_apellido(value: Any) -> str:
    text = to_upper(value)
    if not text:
        raise RowIssue("Apellido vacio")
    return text


def norm_nombre(value: Any) -> str:
    text = to_title_case(value)
    if not text:
        raise RowIssue("Nombre vacio")
    return text


def norm_choice(raw: Any, mapper: dict[str, str], allowed: set[str], field_name: str) -> str:
    text = norm_spaces(raw)
    if not text:
        return ""
    key = text.lower()
    mapped = mapper.get(key, text)
    if mapped not in allowed:
        raise RowIssue(f"Valor invalido para {field_name}: '{raw}'")
    return mapped


def norm_country(raw: Any, field_name: str) -> str:
    allowed = {"Argentina", "Bolivia", "Brasil", "Chile", "Paraguay", "Uruguay", "Otro"}
    text = norm_spaces(raw)
    if not text:
        return ""
    if text in allowed:
        return text
    # Se registra como "Otro" para no descartar la fila.
    return "Otro"

def normalize_country_with_other(raw: Any) -> tuple[str, str]:
    return normalize_country_with_other_util(raw)


def norm_text(value: Any) -> str:
    return to_title_case(value)


def norm_sexo(raw: Any) -> str:
    text = normalize_sexo(raw)
    if not text:
        return ""
    if text not in {"Masculino", "Femenino", "Otro"}:
        raise RowIssue(f"Valor invalido para sexo: '{raw}'")
    return text


def norm_nivel(raw: Any) -> str:
    allowed = {
        "Primaria Completa",
        "Secundaria Incompleta",
        "Secundaria Completa",
        "Terciaria/Universitaria Incompleta",
        "Terciaria/Universitaria Completa",
        "Terciaria/Universitaria",
    }
    text = norm_spaces(raw)
    if not text:
        return ""
    mapped = NIVEL_MAP.get(text.lower(), text)
    if mapped not in allowed:
        raise RowIssue(f"Nivel educativo invalido: '{raw}'")
    return mapped


def norm_bool(raw: Any, field_name: str) -> bool:
    text = norm_spaces(raw).lower()
    if text in TRUE_SET:
        return True
    if text in FALSE_SET or text == "":
        return False
    raise RowIssue(f"Booleano invalido para {field_name}: '{raw}'")


def norm_phone(raw: Any) -> str:
    digits = re.sub(r"\D", "", norm_spaces(raw))
    if not digits:
        return ""
    if len(digits) != 10:
        raise RowIssue(f"Telefono invalido '{raw}'. Debe tener 10 digitos.")
    return digits


def norm_date(raw: Any) -> date | None:
    if raw in (None, ""):
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw

    text = norm_spaces(raw)
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    raise RowIssue(f"Fecha invalida '{raw}'")


def norm_timestamp(raw: Any) -> datetime:
    if raw in (None, ""):
        raise RowIssue("Marca temporal vacia")
    if isinstance(raw, datetime):
        return raw
    if isinstance(raw, date):
        return datetime.combine(raw, datetime.min.time())

    text = norm_spaces(raw)
    for fmt in (
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%d/%m/%Y",
        "%Y-%m-%d",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue

    raise RowIssue(f"Marca temporal invalida '{raw}'")


def write_logs(entries: list[dict[str, Any]], summary: dict[str, Any], source_file: Path) -> Path:
    logs_dir = source_file.resolve().parent / "logs"
    logs_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = logs_dir / f"carga_estudiantes_{ts}.jsonl"

    with log_path.open("w", encoding="utf-8") as f:
        f.write(json.dumps({"type": "summary", **summary}, ensure_ascii=False) + "\n")
        for e in entries:
            f.write(json.dumps(e, ensure_ascii=False) + "\n")

    return log_path


def main() -> int:
    parser = argparse.ArgumentParser(description="Carga masiva de estudiantes desde XLSX")
    parser.add_argument("--file", required=True, help="Ruta al archivo xlsx")
    parser.add_argument("--sheet", default=None, help="Nombre de hoja (default: activa)")
    parser.add_argument("--dry-run", action="store_true", help="Solo valida y simula, no guarda")
    args = parser.parse_args()

    input_path = Path(args.file)
    if not input_path.exists():
        print(f"ERROR: no existe el archivo {input_path}")
        return 1

    wb = load_workbook(filename=input_path, data_only=True, read_only=True)
    ws = wb[args.sheet] if args.sheet else wb.active

    rows = ws.iter_rows(values_only=True)
    headers = [norm_spaces(h) for h in next(rows)]

    missing = [h for h in EXPECTED_HEADERS if h not in headers]
    if missing:
        print("ERROR: faltan encabezados esperados:")
        for m in missing:
            print(f"- {m}")
        return 1

    idx = {h: headers.index(h) for h in EXPECTED_HEADERS}

    issues: list[dict[str, Any]] = []
    total = 0
    created = 0
    updated = 0
    unchanged = 0
    errors = 0
    duplicates_discarded = 0

    selected_rows_by_dni: dict[str, dict[str, Any]] = {}

    for excel_row_num, row in enumerate(rows, start=2):
        if row is None:
            continue

        # Fila vacia completa
        if all(norm_spaces(v) == "" for v in row):
            continue

        total += 1
        try:
            marca_temporal = norm_timestamp(row[idx["Marca temporal"]])
            dni = norm_dni(row[idx["DNI"]])
            current = selected_rows_by_dni.get(dni)
            if current is None or marca_temporal >= current["marca_temporal"]:
                if current is not None:
                    duplicates_discarded += 1
                    issues.append(
                        {
                            "type": "warning",
                            "row": current["row_num"],
                            "dni_raw": dni,
                            "detail": f"Fila descartada por existir otra mas reciente para DNI {dni}",
                        }
                    )
                selected_rows_by_dni[dni] = {
                    "row_num": excel_row_num,
                    "row": row,
                    "marca_temporal": marca_temporal,
                }
            else:
                duplicates_discarded += 1
                issues.append(
                    {
                        "type": "warning",
                        "row": excel_row_num,
                        "dni_raw": dni,
                        "detail": f"Fila descartada por ser mas antigua que otra para DNI {dni}",
                    }
                )

        except (RowIssue, ValidationError) as exc:
            errors += 1
            detail = exc.message_dict if isinstance(exc, ValidationError) else str(exc)
            issues.append(
                {
                    "type": "error",
                    "row": excel_row_num,
                    "dni_raw": norm_spaces(row[idx["DNI"]]) if idx.get("DNI") is not None else "",
                    "email_raw": norm_spaces(row[idx["Correo Electrónico"]]) if idx.get("Correo Electrónico") is not None else "",
                    "detail": detail,
                }
            )
    for item in selected_rows_by_dni.values():
        row = item["row"]
        excel_row_num = item["row_num"]
        marca_temporal = item["marca_temporal"]
        try:
            email = norm_email(row[idx["Correo Electrónico"]])
            apellido = norm_apellido(row[idx["Apellidos"]])
            nombre = norm_nombre(row[idx["Nombres"]])
            dni = norm_dni(row[idx["DNI"]])

            if not email:
                raise RowIssue("Correo Electronico vacio")

            sexo = norm_sexo(row[idx["Sexo"]])
            fecha_nacimiento = norm_date(row[idx["Fecha de Nacimiento"]])
            pais_nacimiento, pais_nacimiento_otro = normalize_country_with_other(row[idx["País de Nacimiento"]])
            nacionalidad, nacionalidad_otra = normalize_country_with_other(row[idx["Nacionalidad"]])
            telefono = norm_phone(row[idx["Teléfono (2964123456)"]])
            nivel_educativo = norm_nivel(row[idx["Estudios de nivel alcanzado"]])

            payload = {
                "email": email,
                "apellido": apellido,
                "nombre": nombre,
                "cuit": norm_cuit(row[idx["CUIL"]]),
                "sexo": sexo,
                "fecha_nacimiento": fecha_nacimiento,
                "pais_nacimiento": pais_nacimiento,
                "pais_nacimiento_otro": pais_nacimiento_otro,
                "nacionalidad": nacionalidad,
                "nacionalidad_otra": nacionalidad_otra,
                "lugar_nacimiento": norm_text(row[idx["Lugar de Nacimiento"]]),
                "domicilio": norm_text(row[idx["Domicilio Particular"]]),
                "barrio": norm_text(row[idx["Barrio"]]),
                "ciudad": norm_text(row[idx["Ciudad"]]),
                "telefono": telefono,
                "nivel_educativo": nivel_educativo,
                "posee_pc": norm_bool(row[idx["¿Posee pc en su domicilio?"]], "posee_pc"),
                "posee_conectividad": norm_bool(
                    row[idx["¿Posee conectividad a internet?"]], "posee_conectividad"
                ),
                "puede_traer_pc": norm_bool(
                    row[idx["¿Puede asistir a clase con su computadora personal?"]], "puede_traer_pc"
                ),
                "trabaja": norm_bool(row[idx["¿Trabaja?"]], "trabaja"),
                "lugar_trabajo": norm_text(row[idx["Lugar de trabajo"]]),
            }

            if not payload["trabaja"]:
                payload["lugar_trabajo"] = ""

            if Estudiante.objects.filter(email=email).exclude(dni=dni).exists():
                raise RowIssue(f"Email duplicado con otro DNI: {email}")

            if args.dry_run:
                existente = Estudiante.objects.filter(dni=dni).first()
                if existente is None:
                    created += 1
                else:
                    changed_fields = [k for k, v in payload.items() if getattr(existente, k) != v]
                    if changed_fields:
                        updated += 1
                    else:
                        unchanged += 1
                continue

            with transaction.atomic():
                estudiante, was_created = Estudiante.objects.get_or_create(dni=dni, defaults=payload)
                if was_created:
                    estudiante.full_clean()
                    estudiante.save()
                    created += 1
                else:
                    changed_fields = []
                    for field, value in payload.items():
                        if getattr(estudiante, field) != value:
                            setattr(estudiante, field, value)
                            changed_fields.append(field)
                    if changed_fields:
                        estudiante.full_clean()
                        estudiante.save(update_fields=changed_fields + ["updated_at"])
                        updated += 1
                    else:
                        unchanged += 1

                marca_temporal_tz = marca_temporal
                if timezone.is_naive(marca_temporal_tz):
                    marca_temporal_tz = timezone.make_aware(
                        marca_temporal_tz, timezone.get_current_timezone()
                    )
                Estudiante.objects.filter(pk=estudiante.pk).update(created_at=marca_temporal_tz)

        except (RowIssue, ValidationError) as exc:
            errors += 1
            detail = exc.message_dict if isinstance(exc, ValidationError) else str(exc)
            issues.append(
                {
                    "type": "error",
                    "row": excel_row_num,
                    "dni_raw": norm_spaces(row[idx["DNI"]]) if idx.get("DNI") is not None else "",
                    "email_raw": norm_spaces(row[idx["Correo Electrónico"]]) if idx.get("Correo Electrónico") is not None else "",
                    "detail": detail,
                }
            )

    summary = {
        "dry_run": args.dry_run,
        "file": str(input_path),
        "sheet": ws.title,
        "total_rows_processed": total,
        "created": created,
        "updated": updated,
        "unchanged": unchanged,
        "errors": errors,
        "duplicates_discarded": duplicates_discarded,
        "timestamp": datetime.now().isoformat(),
    }

    log_path = write_logs(issues, summary, input_path)

    print("=" * 60)
    print("RESUMEN CARGA ESTUDIANTES")
    print("=" * 60)
    for k, v in summary.items():
        print(f"{k}: {v}")
    print(f"log_file: {log_path}")
    print("=" * 60)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
